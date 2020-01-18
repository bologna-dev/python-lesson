#
# consume_sum_g
#
import openpyxl
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

#  Excelファイル選択ダイアログ表示、ファイルパス取得
def OpenFileDlg(tbox):
   ftype = [('','*')] 
   dir = '.'
   # ファイルダイアログ表示
   filename = filedialog.askopenfilename(filetypes = ftype, initialdir = dir)
   # ファイルパスをテキストボックスに表示
   tbox.insert(0, filename)

# Excelデータ集計
def ExcelCalc(fpath):
   # 定数
   SHEET_NAME = 'Sheet1' # シート名
   TYPE_COL = 3  # 入金/出金
   AMOUNT_COL = 4  # 金額カラム
   DATA_START_ROW = 3 # データ開始行
   DATA_END_ROW = 10 # データ終了行
   SHUKIN_LABEL = '出金'
   NYUKIN_LABEL = '入金'

   # エクセルファイル読み込み
   wb = openpyxl.load_workbook(fpath)
   # シート読み込み
   sheet = wb[SHEET_NAME]

   i = 0
   shukin = 0    # 出金額
   nyukin = 0    # 入金額
   # 各行のデータを読み込み
   for i in range(DATA_START_ROW,DATA_END_ROW):
      # 出金だったら
      if sheet.cell(row=i, column=TYPE_COL).value == SHUKIN_LABEL:
         # 出金額に加える
         shukin += sheet.cell(row=i, column=AMOUNT_COL).value
      # 入金だったら
      elif sheet.cell(row=i, column=TYPE_COL).value == NYUKIN_LABEL:
         # 入金額に加える
         nyukin += sheet.cell(row=i, column=AMOUNT_COL).value
   
   return nyukin, shukin

# 集計コールバック
def CalcSum(fpath, nvar, svar):
   # Excelファイルが指定されていない場合、エラー表示
   if fpath == '':
      messagebox.showerror('エラー','Excelファイルを指定してください')
      return
   # 集計
   results = ExcelCalc(fpath)
   # 結果表示
   nvar.set(results[0])
   svar.set(results[1])

# 閉じるコールバック
def DoExit():
   exit()

#
# メイン
#
root = tk.Tk()
root.title('Excel集計')
root.geometry("370x170")
# Excelファイルダイアログ
label = tk.Label(root, text='Excelファイル')
label.place(x=30, y=10)
file_text = tk.Entry(root, width=40)
file_text.place(x=30, y=30)
fdlg_button = tk.Button(root, text='ファイル選択', command = lambda: OpenFileDlg(file_text) )
fdlg_button.place(x=280, y=30)
# 集計表示
nyukin_label = tk.Label(root, text='入金合計：')
nyukin_label.place(x=30, y=120)
nyukin_var = tk.StringVar()
nyukin_var.set("0")
nyukin_value = tk.Label(root, textvariable=nyukin_var)
nyukin_value.place(x=100, y=120)
shukin_label = tk.Label(root, text='出金合計：')
shukin_label.place(x=30, y=140)
shukin_var = tk.StringVar()
shukin_var.set("0")
shukin_value = tk.Label(root, textvariable=shukin_var)
shukin_value.place(x=100, y=140)
# 集計ボタン
calc_button = tk.Button(root, text='集  計', command = lambda: CalcSum(file_text.get(), nyukin_var, shukin_var))
calc_button.place(x=80, y=80)
# 閉じるボタン
close_button = tk.Button(root, text='閉じる', command = lambda: DoExit())
close_button.place(x=200, y=80)
root.mainloop()