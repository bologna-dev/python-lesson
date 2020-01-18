# Excelライブラリ読み込み
import openpyxl

# エクセルファイル読み込み
wb = openpyxl.load_workbook('kakeibo.xlsx')
# シート読み込み
sheet = wb['Sheet1']

i = 0
shukin = 0    # 出金額
nyukin = 0    # 入金額
# 各行のデータを読み込み
for i in range(3,10):
    # 出金だったら
    if sheet.cell(row=i, column=3).value == '出金':
        # 出金額に加える
        shukin += sheet.cell(row=i, column=4).value
    # 入金だったら
    elif sheet.cell(row=i, column=3).value == '入金':
        # 入金額に加える
        nyukin += sheet.cell(row=i, column=4).value
# 出力 
print("出金額：" + str(shukin))
print("入金額：" + str(nyukin))